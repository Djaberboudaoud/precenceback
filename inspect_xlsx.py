"""
Read centres.xlsx and print all rows so we can see the structure.
"""
import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\Dell\Downloads\7odor\backend\centres.xlsx")
ws = wb.active

print(f"Columns: {[cell.value for cell in ws[1]]}")
print(f"Total rows (incl. header): {ws.max_row}\n")

for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(row)
    if i > 5:        # print first 6 rows then stop
        print("...")
        break
