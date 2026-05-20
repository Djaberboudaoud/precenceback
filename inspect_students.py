import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\Dell\Downloads\7odor\backend\new_convocation-2026.xlsx")
ws = wb.active

print(f"Columns: {[cell.value for cell in ws[1]]}")

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i > 0:
        print(row)
    if i > 5:
        break
